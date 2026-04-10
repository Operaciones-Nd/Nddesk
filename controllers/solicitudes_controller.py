from flask import Blueprint, render_template, request, redirect, url_for, flash, session, jsonify, send_file, current_app, abort
from werkzeug.utils import secure_filename
from models import Solicitud, Usuario, ComentarioTicket, AdjuntoTicket, CalificacionTicket, TipoTicket, SLATracking, Escalamiento, db
from services import TicketService, SLAService, EscalamientoService
from services.transicion_service import TransicionService
from utils import login_required, ValidationError, get_client_ip
from utils.file_validator import FileValidator
from utils.validators import InputSanitizer
from utils.rate_limit import rate_limit
from datetime import datetime, timedelta
from sqlalchemy import func
from pathlib import Path
import os

solicitudes_bp = Blueprint('solicitudes', __name__)

ALLOWED_EXTENSIONS = {'pdf', 'png', 'jpg', 'jpeg', 'doc', 'docx', 'xls', 'xlsx', 'txt'}
MAX_FILE_SIZE = 10 * 1024 * 1024  # 10MB

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@solicitudes_bp.route('/estadisticas/exportar-pdf')
@login_required
def exportar_estadisticas_pdf():
    import pdfkit
    from io import BytesIO
    from datetime import datetime
    
    # Renderizar la página HTML
    html_content = render_template('solicitudes/estadisticas_pdf.html', 
                                   metricas=get_metricas(), 
                                   stats=get_stats())
    
    # Configuración de wkhtmltopdf
    options = {
        'page-size': 'A4',
        'margin-top': '0.5in',
        'margin-right': '0.5in',
        'margin-bottom': '0.5in',
        'margin-left': '0.5in',
        'encoding': 'UTF-8',
        'no-outline': None,
        'enable-local-file-access': None,
        'javascript-delay': 2000
    }
    
    # Generar PDF
    pdf = pdfkit.from_string(html_content, False, options=options)
    
    buffer = BytesIO(pdf)
    buffer.seek(0)
    
    return send_file(
        buffer,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=f'estadisticas_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
    )

def get_metricas():
    return {
        'pendientes': Solicitud.query.filter_by(estado='Pendiente', deleted_at=None).count(),
        'planificados': Solicitud.query.filter_by(estado='Planificado', deleted_at=None).count(),
        'solucionados': Solicitud.query.filter_by(estado='Solucionado', deleted_at=None).count(),
        'cerrados': Solicitud.query.filter_by(estado='Cerrado', deleted_at=None).count(),
        'escalados': Escalamiento.query.count(),
        'sla_vencido': SLATracking.query.filter(SLATracking.resolucion_cumplida == False).count()
    }

def get_stats():
    from datetime import timedelta
    
    total_tickets = Solicitud.query.filter_by(deleted_at=None).count()
    tickets_resueltos = Solicitud.query.filter(Solicitud.estado.in_(['Solucionado', 'Cerrado']), Solicitud.deleted_at == None).count()
    
    prioridad_alta = Solicitud.query.filter_by(prioridad='Alta', deleted_at=None).count()
    prioridad_media = Solicitud.query.filter_by(prioridad='Media', deleted_at=None).count()
    prioridad_baja = Solicitud.query.filter_by(prioridad='Baja', deleted_at=None).count()
    
    tickets_por_seccion = db.session.query(
        Solicitud.seccion,
        func.count(Solicitud.id)
    ).filter(Solicitud.deleted_at == None).group_by(Solicitud.seccion).order_by(func.count(Solicitud.id).desc()).all()
    
    tickets_por_departamento = db.session.query(
        Solicitud.departamento,
        func.count(Solicitud.id)
    ).filter(Solicitud.deleted_at == None).group_by(Solicitud.departamento).order_by(func.count(Solicitud.id).desc()).all()
    
    hoy = datetime.now().date()
    tendencia_7dias = []
    labels_7dias = []
    for i in range(6, -1, -1):
        fecha = hoy - timedelta(days=i)
        count = Solicitud.query.filter(
            func.date(Solicitud.created_at) == fecha,
            Solicitud.deleted_at == None
        ).count()
        tendencia_7dias.append(count)
        if i == 0:
            labels_7dias.append('Hoy')
        elif i == 1:
            labels_7dias.append('Ayer')
        else:
            labels_7dias.append(f'Hace {i}d')
    
    tickets_cerrados = Solicitud.query.filter(
        Solicitud.estado == 'Cerrado',
        Solicitud.fecha_cierre != None,
        Solicitud.deleted_at == None
    ).all()
    
    if tickets_cerrados:
        tiempos = []
        for ticket in tickets_cerrados:
            if ticket.created_at and ticket.fecha_cierre:
                diff = ticket.fecha_cierre - ticket.created_at
                tiempos.append(diff.total_seconds() / 3600)
        tiempo_promedio = round(sum(tiempos) / len(tiempos), 1) if tiempos else 0
    else:
        tiempo_promedio = 0
    
    tasa_resolucion = round((tickets_resueltos / total_tickets * 100) if total_tickets > 0 else 0, 1)
    
    calificaciones = CalificacionTicket.query.limit(1000).all()
    if calificaciones:
        satisfaccion = round(sum([c.calificacion for c in calificaciones]) / len(calificaciones), 1)
    else:
        satisfaccion = 0
    
    top_resolutores = db.session.query(
        Usuario.nombre,
        func.count(Solicitud.id).label('total')
    ).join(
        Solicitud, Solicitud.resuelto_por == Usuario.id
    ).filter(
        Solicitud.estado.in_(['Solucionado', 'Cerrado']),
        Solicitud.deleted_at == None
    ).group_by(Usuario.id).order_by(func.count(Solicitud.id).desc()).limit(5).all()
    
    return {
        'tiempo_promedio': tiempo_promedio,
        'tasa_resolucion': tasa_resolucion,
        'satisfaccion': satisfaccion,
        'prioridad_alta': prioridad_alta,
        'prioridad_media': prioridad_media,
        'prioridad_baja': prioridad_baja,
        'total_tickets': total_tickets,
        'tickets_resueltos': tickets_resueltos,
        'tickets_por_seccion': dict(tickets_por_seccion),
        'tickets_por_departamento': dict(tickets_por_departamento),
        'tendencia_7dias': tendencia_7dias,
        'labels_7dias': labels_7dias,
        'top_resolutores': top_resolutores
    }

@solicitudes_bp.route('/estadisticas')
@login_required
def estadisticas():
    from sqlalchemy import func, extract
    from datetime import datetime, timedelta
    
    user = Usuario.query.get(session['user_id'])
    
    # Query base según rol
    if user.is_admin or user.is_coordinador:
        base_query = Solicitud.query.filter_by(deleted_at=None)
    elif user.is_resolutor:
        secciones_usuario = [s.nombre for s in user.secciones]
        base_query = Solicitud.query.filter_by(deleted_at=None).filter(
            db.or_(
                Solicitud.seccion.in_(secciones_usuario),
                Solicitud.resuelto_por == user.id
            )
        )
    else:
        # Solicitantes ven estadísticas de su departamento
        if user.departamento_obj:
            base_query = Solicitud.query.filter_by(deleted_at=None, departamento=user.departamento_obj.nombre)
        else:
            base_query = Solicitud.query.filter_by(deleted_at=None, usuario_id=user.id)
    
    # Métricas básicas
    metricas = {
        'pendientes': base_query.filter_by(estado='Pendiente').count(),
        'planificados': base_query.filter_by(estado='Planificado').count(),
        'solucionados': base_query.filter_by(estado='Solucionado').count(),
        'cerrados': base_query.filter_by(estado='Cerrado').count(),
        'escalados': base_query.filter_by(estado='Escalado').count() if user.is_admin or user.is_coordinador else 0,
        'sla_vencido': 0
    }
    
    # Total de tickets
    total_tickets = base_query.count()
    tickets_resueltos = metricas['solucionados'] + metricas['cerrados']
    
    # Tickets por prioridad
    prioridad_alta = base_query.filter_by(prioridad='Alta').count()
    prioridad_media = base_query.filter_by(prioridad='Media').count()
    prioridad_baja = base_query.filter_by(prioridad='Baja').count()
    
    # Tickets por sección
    tickets_por_seccion = db.session.query(
        Solicitud.seccion,
        func.count(Solicitud.id)
    ).filter(
        Solicitud.deleted_at == None
    ).group_by(Solicitud.seccion).order_by(func.count(Solicitud.id).desc()).all()
    
    # Tickets por departamento
    tickets_por_departamento = db.session.query(
        Solicitud.departamento,
        func.count(Solicitud.id)
    ).filter(
        Solicitud.deleted_at == None
    ).group_by(Solicitud.departamento).order_by(func.count(Solicitud.id).desc()).all()
    
    # Tickets por tipo de contenido
    tickets_por_tipo = db.session.query(
        Solicitud.tipo_contenido,
        func.count(Solicitud.id)
    ).filter(
        Solicitud.id.in_([t.id for t in base_query.all()])
    ).group_by(Solicitud.tipo_contenido).all()
    
    # Tickets por familia de servicios
    tickets_por_servicio = db.session.query(
        Solicitud.familia_servicios,
        func.count(Solicitud.id)
    ).filter(
        Solicitud.id.in_([t.id for t in base_query.all()]),
        Solicitud.familia_servicios != None
    ).group_by(Solicitud.familia_servicios).all()
    
    # Tendencia últimos 7 días
    hoy = datetime.now().date()
    tendencia_7dias = []
    labels_7dias = []
    all_tickets = base_query.all()
    for i in range(6, -1, -1):
        fecha = hoy - timedelta(days=i)
        count = sum(1 for t in all_tickets if t.created_at.date() == fecha)
        tendencia_7dias.append(count)
        if i == 0:
            labels_7dias.append('Hoy')
        elif i == 1:
            labels_7dias.append('Ayer')
        else:
            labels_7dias.append(f'Hace {i}d')
    
    # Tiempo promedio de resolución (optimizado)
    tiempo_promedio_query = db.session.query(
        func.avg(
            func.julianday(Solicitud.fecha_cierre) - func.julianday(Solicitud.created_at)
        ) * 24
    ).filter(
        Solicitud.id.in_([t.id for t in base_query.all()]),
        Solicitud.estado == 'Cerrado',
        Solicitud.fecha_cierre != None,
        Solicitud.created_at != None
    ).scalar()
    
    tiempo_promedio = round(tiempo_promedio_query, 1) if tiempo_promedio_query else 0
    
    # Tasa de resolución
    tasa_resolucion = round((tickets_resueltos / total_tickets * 100) if total_tickets > 0 else 0, 1)
    
    # Satisfacción promedio
    calificaciones = CalificacionTicket.query.limit(1000).all()
    if calificaciones:
        satisfaccion = round(sum([c.calificacion for c in calificaciones]) / len(calificaciones), 1)
    else:
        satisfaccion = 0
    
    # Top 5 resolutores
    top_resolutores = db.session.query(
        Usuario.nombre,
        func.count(Solicitud.id).label('total')
    ).join(
        Solicitud, Solicitud.resuelto_por == Usuario.id
    ).filter(
        Solicitud.estado.in_(['Solucionado', 'Cerrado']),
        Solicitud.deleted_at == None
    ).group_by(Usuario.id).order_by(func.count(Solicitud.id).desc()).limit(5).all()
    
    stats = {
        'tiempo_promedio': tiempo_promedio,
        'tasa_resolucion': tasa_resolucion,
        'satisfaccion': satisfaccion,
        'prioridad_alta': prioridad_alta,
        'prioridad_media': prioridad_media,
        'prioridad_baja': prioridad_baja,
        'total_tickets': total_tickets,
        'tickets_resueltos': tickets_resueltos,
        'tickets_por_seccion': dict(tickets_por_seccion),
        'tickets_por_departamento': dict(tickets_por_departamento),
        'tickets_por_tipo': dict(tickets_por_tipo),
        'por_servicio': dict(tickets_por_servicio),
        'tendencia_7dias': tendencia_7dias,
        'labels_7dias': labels_7dias,
        'top_resolutores': top_resolutores
    }
    
    return render_template('solicitudes/estadisticas.html', metricas=metricas, stats=stats)

@solicitudes_bp.route('/')
@login_required
def index():
    from models import Seccion
    user = Usuario.query.get(session['user_id'])
    
    # Métricas (optimizadas con queries directas)
    fecha_30_dias = datetime.now() - timedelta(days=30)
    
    # Base query según rol
    if user.is_admin or user.is_coordinador:
        base_query = Solicitud.query.filter_by(deleted_at=None)
    elif user.is_resolutor:
        secciones_usuario = [s.nombre for s in user.secciones]
        base_query = Solicitud.query.filter_by(deleted_at=None).filter(
            db.or_(
                Solicitud.seccion.in_(secciones_usuario),
                Solicitud.resuelto_por == user.id
            )
        )
    else:
        # Solicitantes ven tickets de su departamento
        if user.departamento_obj:
            base_query = Solicitud.query.filter_by(deleted_at=None, departamento=user.departamento_obj.nombre)
        else:
            base_query = Solicitud.query.filter_by(deleted_at=None, usuario_id=user.id)
    
    metricas = {
        'pendientes': base_query.filter_by(estado='Pendiente').count(),
        'planificados': base_query.filter_by(estado='Planificado').count(),
        'solucionados': base_query.filter_by(estado='Solucionado').count(),
        'cerrados': base_query.filter_by(estado='Cerrado').count(),
        'escalados': base_query.filter_by(estado='Escalado').count(),
        'activos': base_query.filter(
            Solicitud.estado.in_(['Pendiente', 'Planificado', 'Escalado']),
            Solicitud.created_at >= fecha_30_dias
        ).count()
    }
    
    # Obtener secciones para filtros (solo activas)
    from models import Seccion
    secciones = [s.nombre for s in Seccion.get_active().all()]
    
    # Prioridades
    prioridades = ['Alta', 'Media', 'Baja']
    
    # Tickets recientes (últimos 10)
    tickets_recientes = base_query.order_by(Solicitud.created_at.desc()).limit(10).all()
    
    return render_template('solicitudes/index.html', metricas=metricas, secciones=secciones, prioridades=prioridades, tickets_recientes=tickets_recientes)

@solicitudes_bp.route('/nueva', methods=['GET', 'POST'])
@login_required
def nueva():
    from models import Seccion, Departamento, Servicio, Subcategoria
    
    user = Usuario.query.get(session['user_id'])
    
    if request.method == 'POST':
        try:
            # Validar duplicados (mismo usuario, misma descripción, últimos 5 minutos)
            hace_5_min = datetime.now() - timedelta(minutes=5)
            duplicado = Solicitud.query.filter(
                Solicitud.usuario_id == user.id,
                Solicitud.descripcion == request.form['descripcion'],
                Solicitud.created_at >= hace_5_min,
                Solicitud.deleted_at == None
            ).first()
            
            if duplicado:
                flash(f'Ya creaste un ticket similar hace poco (#{duplicado.id})', 'warning')
                return redirect(url_for('solicitudes.ver', id=duplicado.id))
            
            seccion = request.form['seccion']
            grupo_resuelve = request.form.get('grupo_resuelve', seccion)
            
            # Buscar email de notificación de la sección
            from models import Seccion
            seccion_obj = Seccion.query.filter_by(nombre=seccion, deleted_at=None).first()
            email_notificacion = seccion_obj.email_notificacion if seccion_obj and seccion_obj.email_notificacion else request.form.get('email_notificacion', '')
            
            # NO asignar resolutor automáticamente - el equipo lo toma manualmente
            
            data = {
                'fecha_publicacion': datetime.strptime(request.form['fecha_publicacion'], '%Y-%m-%d').date(),
                'medio': request.form['medio'],
                'departamento': request.form.get('departamento') if not user.departamento_id else (user.departamento_obj.nombre if user.departamento_obj else 'Sin departamento'),
                'seccion': seccion,
                'familia_servicios': request.form['familia_servicios'],
                'servicio': request.form.get('servicio', seccion),
                'grupo_resuelve': grupo_resuelve,
                'email_notificacion': email_notificacion,
                'tipo_contenido': request.form['tipo_contenido'],
                'descripcion': InputSanitizer.sanitize_string(request.form['descripcion'], max_length=5000),
                'prioridad': request.form.get('prioridad', 'Media')
            }
            
            # Validar que la fecha no sea anterior a hoy
            fecha_pub = data['fecha_publicacion']
            if fecha_pub < datetime.now().date():
                flash('No se pueden crear publicaciones en fechas anteriores', 'danger')
                return redirect(url_for('solicitudes.nueva'))
            
            ticket = TicketService.create_ticket(data, user, get_client_ip(request))
            
            flash(f'Ticket #{ticket.id} creado y asignado al grupo {grupo_resuelve}', 'success')
            
            return redirect(url_for('solicitudes.ver', id=ticket.id))
        except ValidationError as e:
            flash(str(e), 'danger')
        except Exception as e:
            flash(f'Error al crear ticket: {str(e)}', 'danger')
    
    # Cargar datos dinámicos
    secciones = Seccion.query.filter_by(deleted_at=None, activo=True).all()
    departamentos = Departamento.query.filter_by(deleted_at=None, activo=True).all()
    servicios = Servicio.query.filter_by(deleted_at=None, activo=True).all()
    today = datetime.now().strftime('%Y-%m-%d')
    
    return render_template('solicitudes/nueva.html', user=user, secciones=secciones, departamentos=departamentos, servicios=servicios, today=today)

@solicitudes_bp.route('/<int:id>')
@login_required
def ver(id):
    from models import Usuario
    ticket = Solicitud.query.get_or_404(id)
    comentarios = ComentarioTicket.get_by_ticket(id)
    adjuntos = AdjuntoTicket.get_by_ticket(id)
    calificacion = CalificacionTicket.get_by_ticket(id)
    resolutores = Usuario.query.filter(
        Usuario.rol.in_(['resolutor', 'coordinador', 'admin']),
        Usuario.activo == True,
        Usuario.deleted_at == None
    ).limit(100).all()
    return render_template('solicitudes/ver_nuevo.html', ticket=ticket, comentarios=comentarios, adjuntos=adjuntos, calificacion=calificacion, resolutores=resolutores)

@solicitudes_bp.route('/<int:id>/editar', methods=['GET', 'POST'])
@login_required
def editar(id):
    ticket = Solicitud.query.get_or_404(id)
    user = Usuario.query.get(session['user_id'])
    
    # Validación estricta de permisos
    puede_editar = (
        user.is_admin or 
        user.is_coordinador or 
        (user.is_resolutor and ticket.resuelto_por == user.id)
    )
    
    if not puede_editar:
        current_app.logger.warning(f'Intento de edición no autorizada: usuario {user.id} ticket {id}')
        flash('No tienes permisos para editar este ticket', 'danger')
        return redirect(url_for('solicitudes.ver', id=id))
    
    if ticket.estado == 'Cerrado' and not (user.is_admin or user.is_coordinador):
        flash('No puedes editar un ticket cerrado', 'danger')
        return redirect(url_for('solicitudes.ver', id=id))
    
    if request.method == 'POST':
        try:
            data = {
                'estado': request.form.get('estado'),
                'prioridad': request.form.get('prioridad'),
                'solucion': request.form.get('solucion'),
                'bitacora_publica': request.form.get('bitacora_publica'),
                'bitacora_oculta': request.form.get('bitacora_oculta')
            }
            
            TicketService.update_ticket(ticket, data, user, get_client_ip(request))
            flash('Ticket actualizado', 'success')
            return redirect(url_for('solicitudes.ver', id=ticket.id))
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'danger')
    
    return render_template('solicitudes/editar.html', ticket=ticket)

@solicitudes_bp.route('/<int:id>/tomar-ticket', methods=['POST'])
@login_required
def tomar_ticket(id):
    from models import AuditoriaTicket
    ticket = Solicitud.query.get_or_404(id)
    user = Usuario.query.get(session['user_id'])
    
    if not (user.is_admin or user.is_coordinador or user.is_resolutor):
        flash('No tienes permisos', 'danger')
        return redirect(url_for('solicitudes.ver', id=id))
    
    if ticket.resuelto_por:
        flash('Este ticket ya está asignado', 'warning')
        return redirect(url_for('solicitudes.ver', id=id))
    
    try:
        ticket.resuelto_por = user.id
        
        # Auditoría
        auditoria = AuditoriaTicket(
            ticket_id=ticket.id,
            usuario_id=user.id,
            accion='tomar_ticket',
            campo_modificado='resuelto_por',
            valor_anterior=None,
            valor_nuevo=str(user.id),
            ip_address=get_client_ip(request)
        )
        db.session.add(auditoria)
        
        db.session.commit()
        flash(f'Ticket asignado a ti exitosamente', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'danger')
    
    return redirect(url_for('solicitudes.ver', id=id))

@solicitudes_bp.route('/<int:id>/marcar-pendiente', methods=['POST'])
@login_required
def marcar_pendiente(id):
    from models import AuditoriaTicket
    ticket = Solicitud.query.get_or_404(id)
    user = Usuario.query.get(session['user_id'])
    
    if not (user.is_admin or user.is_coordinador or user.is_resolutor):
        flash('No tienes permisos', 'danger')
        return redirect(url_for('solicitudes.ver', id=id))
    
    try:
        # Validar transición
        TransicionService.validar_transicion(ticket.estado, 'Pendiente')
        
        estado_anterior = ticket.estado
        ticket.estado = 'Pendiente'
        ticket.motivo_pendiente = request.form.get('motivo_pendiente')
        
        # Auditoría
        auditoria = AuditoriaTicket(
            ticket_id=ticket.id,
            usuario_id=user.id,
            accion='marcar_pendiente',
            campo_modificado='estado',
            valor_anterior=estado_anterior,
            valor_nuevo='Pendiente',
            ip_address=get_client_ip(request)
        )
        db.session.add(auditoria)
        
        db.session.commit()
        flash('Ticket marcado como Pendiente', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'danger')
    
    return redirect(url_for('solicitudes.ver', id=id))

@solicitudes_bp.route('/<int:id>/marcar-solucionado', methods=['POST'])
@login_required
def marcar_solucionado(id):
    from models import AuditoriaTicket
    ticket = Solicitud.query.get_or_404(id)
    user = Usuario.query.get(session['user_id'])
    
    if not (user.is_admin or user.is_coordinador or user.is_resolutor):
        flash('No tienes permisos', 'danger')
        return redirect(url_for('solicitudes.ver', id=id))
    
    try:
        # Validar transición
        TransicionService.validar_transicion(ticket.estado, 'Solucionado')
        
        estado_anterior = ticket.estado
        ticket.estado = 'Solucionado'
        ticket.servicio_solucion = request.form.get('servicio_solucion')
        ticket.subcategoria_solucion = request.form.get('subcategoria_solucion')
        ticket.codigo_solucion = request.form.get('codigo_solucion')
        ticket.solucion = request.form.get('solucion')
        ticket.resuelto_por = user.id
        
        # Auditoría
        auditoria = AuditoriaTicket(
            ticket_id=ticket.id,
            usuario_id=user.id,
            accion='marcar_solucionado',
            campo_modificado='estado',
            valor_anterior=estado_anterior,
            valor_nuevo='Solucionado',
            ip_address=get_client_ip(request)
        )
        db.session.add(auditoria)
        
        db.session.commit()
        flash('Ticket marcado como Solucionado', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'danger')
    
    return redirect(url_for('solicitudes.ver', id=id))

@solicitudes_bp.route('/<int:id>/reasignar', methods=['POST'])
@login_required
def reasignar(id):
    from models import AuditoriaTicket
    ticket = Solicitud.query.get_or_404(id)
    user = Usuario.query.get(session['user_id'])
    
    if not (user.is_admin or user.is_coordinador):
        flash('No tienes permisos', 'danger')
        return redirect(url_for('solicitudes.ver', id=id))
    
    try:
        resolutor_anterior = ticket.resuelto_por
        nuevo_resolutor_id = request.form.get('resuelto_por')
        
        ticket.resuelto_por = nuevo_resolutor_id
        
        # Auditoría
        auditoria = AuditoriaTicket(
            ticket_id=ticket.id,
            usuario_id=user.id,
            accion='reasignar',
            campo_modificado='resuelto_por',
            valor_anterior=str(resolutor_anterior) if resolutor_anterior else None,
            valor_nuevo=str(nuevo_resolutor_id) if nuevo_resolutor_id else None,
            ip_address=get_client_ip(request)
        )
        db.session.add(auditoria)
        
        db.session.commit()
        flash('Ticket reasignado exitosamente', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'danger')
    
    return redirect(url_for('solicitudes.ver', id=id))

@solicitudes_bp.route('/<int:id>/cerrar', methods=['POST'])
@login_required
def cerrar(id):
    from models import AuditoriaTicket
    ticket = Solicitud.query.get_or_404(id)
    user = Usuario.query.get(session['user_id'])
    
    if not (user.is_admin or user.is_coordinador or user.is_resolutor):
        flash('No tienes permisos para cerrar tickets', 'danger')
        return redirect(url_for('solicitudes.ver', id=id))
    
    if ticket.estado == 'Cerrado':
        flash('El ticket ya está cerrado', 'warning')
        return redirect(url_for('solicitudes.ver', id=id))
    
    try:
        from sqlalchemy.exc import SQLAlchemyError
        
        # Validar transición
        TransicionService.validar_transicion(ticket.estado, 'Cerrado')
        
        # Transacción atómica
        estado_anterior = ticket.estado
        ticket.estado = 'Cerrado'
        ticket.estado_publicacion = request.form.get('estado_publicacion')
        ticket.comentarios_usuario = request.form.get('comentarios_usuario')
        ticket.fecha_cierre = datetime.now()
        
        # Auditoría
        auditoria = AuditoriaTicket(
            ticket_id=ticket.id,
            usuario_id=user.id,
            accion='cerrar',
            campo_modificado='estado',
            valor_anterior=estado_anterior,
            valor_nuevo='Cerrado',
            ip_address=get_client_ip(request)
        )
        db.session.add(auditoria)
        
        # Commit atómico
        db.session.commit()
        
        # Notificar cierre al grupo (no debe fallar la transacción)
        try:
            from services.email_service import EmailService
            EmailService.notify_ticket_closed(ticket)
        except Exception as email_error:
            current_app.logger.error(f'Error enviando email de cierre ticket {id}: {email_error}')
            # No fallar la transacción por error de email
        
        flash('Ticket cerrado exitosamente', 'success')
        
    except (ValidationError, SQLAlchemyError) as e:
        db.session.rollback()
        current_app.logger.error(f'Error cerrando ticket {id}: {str(e)}')
        flash(str(e) if isinstance(e, ValidationError) else 'Error al cerrar ticket', 'danger')
    
    return redirect(url_for('solicitudes.ver', id=ticket.id))

@solicitudes_bp.route('/reportes')
@login_required
def reportes():
    user = Usuario.query.get(session['user_id'])
    
    if not (user.is_admin or user.is_coordinador):
        flash('No tienes permisos para acceder a reportes', 'danger')
        return redirect(url_for('solicitudes.index'))
    
    from datetime import date, timedelta
    from models import Seccion
    
    # Obtener secciones para filtros
    secciones = [s.nombre for s in Seccion.get_active().all()]
    prioridades = ['Alta', 'Media', 'Baja']
    estados = ['Pendiente', 'Planificado', 'Escalado', 'Solucionado', 'Cerrado']
    
    # Opciones para filtros
    filtros = {
        'departamentos': db.session.query(Solicitud.departamento).filter(Solicitud.deleted_at == None).distinct().all(),
        'secciones': secciones,
        'estados': estados,
        'prioridades': prioridades,
        'grupos': db.session.query(Solicitud.grupo_resuelve).filter(Solicitud.deleted_at == None).distinct().all(),
        'tipos_contenido': db.session.query(Solicitud.tipo_contenido).filter(Solicitud.deleted_at == None).distinct().all()
    }
    
    return render_template('solicitudes/reportes.html', filtros=filtros)

@solicitudes_bp.route('/reportes/exportar')
@login_required
def exportar_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from io import BytesIO
    
    user = Usuario.query.get(session['user_id'])
    if not (user.is_admin or user.is_coordinador):
        flash('No tienes permisos', 'danger')
        return redirect(url_for('solicitudes.index'))
    
    # Obtener filtros desde DataTables
    fecha_inicio = request.args.get('fecha_inicio')
    fecha_fin = request.args.get('fecha_fin')
    departamento = request.args.get('departamento')
    seccion = request.args.get('seccion')
    estado = request.args.get('estado')
    prioridad = request.args.get('prioridad')
    
    query = Solicitud.query.filter_by(deleted_at=None)
    
    if fecha_inicio:
        query = query.filter(Solicitud.created_at >= fecha_inicio)
    if fecha_fin:
        query = query.filter(Solicitud.created_at <= fecha_fin + ' 23:59:59')
    if departamento:
        query = query.filter(Solicitud.departamento == departamento)
    if seccion:
        query = query.filter(Solicitud.seccion == seccion)
    if estado:
        query = query.filter(Solicitud.estado == estado)
    if prioridad:
        query = query.filter(Solicitud.prioridad == prioridad)
    
    tickets = query.order_by(Solicitud.created_at.desc()).all()
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Tickets"
    
    # Estilos
    header_fill = PatternFill(start_color="1e3a8a", end_color="1e3a8a", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = [
        'ID', 'Fecha Creación', 'Fecha Cierre', 'Medio', 'Departamento', 'Sección',
        'Grupo Resuelve', 'Email', 'Tipo Contenido', 'Estado', 'Prioridad',
        'Descripción', 'Solución', 'Bitácora Pública', 'Bitácora Interna',
        'Solicitante', 'Resuelto Por'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Datos
    for row, ticket in enumerate(tickets, 2):
        solicitante = Usuario.query.get(ticket.usuario_id)
        resuelto_por = Usuario.query.get(ticket.resuelto_por) if ticket.resuelto_por else None
        
        data = [
            ticket.id,
            ticket.created_at.strftime('%d/%m/%Y %H:%M') if ticket.created_at else '',
            ticket.fecha_cierre.strftime('%d/%m/%Y %H:%M') if ticket.fecha_cierre else '',
            ticket.medio or '',
            ticket.departamento or '',
            ticket.seccion or '',
            ticket.grupo_resuelve or '',
            ticket.email_notificacion or '',
            ticket.tipo_contenido or '',
            ticket.estado or '',
            ticket.prioridad or '',
            ticket.descripcion or '',
            ticket.solucion or '',
            ticket.bitacora_publica or '',
            ticket.bitacora_oculta or '',
            solicitante.nombre if solicitante else '',
            resuelto_por.nombre if resuelto_por else ''
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = border
            cell.alignment = Alignment(vertical='top', wrap_text=True)
    
    # Ajustar anchos
    column_widths = [8, 18, 18, 12, 15, 15, 18, 25, 15, 12, 10, 40, 40, 40, 40, 20, 20]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # Filtros
    ws.auto_filter.ref = ws.dimensions
    
    # Guardar
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    from datetime import datetime
    filename = f"reporte_tickets_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

@solicitudes_bp.route('/<int:id>/comentario', methods=['POST'])
@login_required
def agregar_comentario(id):
    ticket = Solicitud.query.get_or_404(id)
    user = Usuario.query.get(session['user_id'])
    
    contenido = request.form.get('contenido', '').strip()
    es_interno = request.form.get('es_interno') == 'true'
    
    # Validaciones
    if not contenido:
        flash('El comentario no puede estar vacío', 'danger')
        return redirect(url_for('solicitudes.ver', id=id) + '#seguimiento')
    
    if len(contenido) > 2000:
        flash('El comentario es demasiado largo (máximo 2000 caracteres)', 'danger')
        return redirect(url_for('solicitudes.ver', id=id) + '#seguimiento')
    
    # Sanitizar HTML
    contenido = InputSanitizer.sanitize_string(contenido, max_length=2000)
    
    try:
        comentario = ComentarioTicket(
            solicitud_id=id,
            usuario_id=user.id,
            contenido=contenido,
            es_interno=es_interno
        )
        db.session.add(comentario)
        
        if not ticket.fecha_primera_respuesta and user.id != ticket.usuario_id:
            ticket.fecha_primera_respuesta = datetime.now()
            SLAService.registrar_primera_respuesta(ticket.id)
        
        db.session.commit()
        current_app.logger.info(f'Comentario agregado al ticket {id} por usuario {user.id}')
        
        # Enviar notificación de nuevo comentario
        try:
            from services.email_service import EmailService
            EmailService.notify_new_comment(ticket, comentario)
        except Exception as email_error:
            current_app.logger.error(f'Error enviando notificación de comentario: {email_error}')
        
        flash('Comentario agregado', 'success')
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f'Error agregando comentario: {str(e)}')
        flash(f'Error: {str(e)}', 'danger')
    
    return redirect(url_for('solicitudes.ver', id=id) + '#seguimiento')

@solicitudes_bp.route('/<int:id>/adjunto/<int:adjunto_id>/descargar')
@login_required
def descargar_adjunto(id, adjunto_id):
    """Descarga segura de archivos adjuntos con validación de permisos"""
    ticket = Solicitud.query.get_or_404(id)
    adjunto = AdjuntoTicket.query.get_or_404(adjunto_id)
    user = Usuario.query.get(session['user_id'])
    
    # Validar que el adjunto pertenece al ticket
    if adjunto.solicitud_id != id:
        current_app.logger.warning(f'Intento de acceso a adjunto incorrecto: usuario {user.id}, adjunto {adjunto_id}, ticket {id}')
        abort(404)
    
    # Validar permisos
    puede_ver = (
        user.is_admin or 
        user.is_coordinador or 
        user.is_resolutor or 
        ticket.usuario_id == user.id
    )
    
    if not puede_ver:
        current_app.logger.warning(f'Intento de descarga no autorizada: usuario {user.id}, adjunto {adjunto_id}')
        abort(403)
    
    # Validar que el archivo existe
    if not os.path.exists(adjunto.ruta_archivo):
        current_app.logger.error(f'Archivo no encontrado: {adjunto.ruta_archivo}')
        flash('Archivo no encontrado', 'danger')
        return redirect(url_for('solicitudes.ver', id=id))
    
    # Servir archivo de forma segura
    try:
        return send_file(
            adjunto.ruta_archivo,
            as_attachment=True,
            download_name=adjunto.nombre_archivo,
            mimetype=adjunto.tipo_mime
        )
    except Exception as e:
        current_app.logger.error(f'Error sirviendo archivo: {str(e)}')
        flash('Error al descargar archivo', 'danger')
        return redirect(url_for('solicitudes.ver', id=id))

@solicitudes_bp.route('/<int:id>/adjunto', methods=['POST'])
@login_required
def subir_archivo(id):
    ticket = Solicitud.query.get_or_404(id)
    user = Usuario.query.get(session['user_id'])
    
    if 'archivo' not in request.files:
        flash('No se seleccionó archivo', 'danger')
        return redirect(url_for('solicitudes.ver', id=id))
    
    file = request.files['archivo']
    
    if file.filename == '':
        flash('No se seleccionó archivo', 'danger')
        return redirect(url_for('solicitudes.ver', id=id))
    
    try:
        # Validación exhaustiva del archivo
        original_filename, mime_type = FileValidator.validate_file(file)
        
        # Sanitizar nombre de archivo
        from werkzeug.utils import secure_filename
        safe_filename = secure_filename(original_filename)
        
        # Generar nombre único
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"{timestamp}_{safe_filename}"
        
        # Crear ruta segura fuera de static
        base_upload = Path(current_app.config.get('UPLOAD_FOLDER', 'uploads'))
        ticket_folder = base_upload / str(id)
        ticket_folder.mkdir(parents=True, exist_ok=True)
        
        # Validar que la ruta final está dentro del directorio permitido
        filepath = (ticket_folder / filename).resolve()
        if not str(filepath).startswith(str(base_upload.resolve())):
            current_app.logger.warning(f'Intento de path traversal: {filepath}')
            abort(400)
        
        # Guardar archivo
        file.save(filepath)
        
        # Registrar en BD
        adjunto = AdjuntoTicket(
            solicitud_id=id,
            usuario_id=user.id,
            nombre_archivo=file.filename,
            ruta_archivo=str(filepath),
            tamano=os.path.getsize(filepath),
            tipo_mime=mime_type
        )
        db.session.add(adjunto)
        db.session.commit()
        
        current_app.logger.info(f'Archivo subido: {filename} al ticket {id} por usuario {user.id}')
        flash('Archivo subido exitosamente', 'success')
        
    except ValidationError as e:
        current_app.logger.warning(f'Validación de archivo fallida: {str(e)}')
        flash(str(e), 'danger')
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f'Error subiendo archivo: {str(e)}')
        flash('Error al subir archivo', 'danger')
    
    return redirect(url_for('solicitudes.ver', id=id))

@solicitudes_bp.route('/<int:id>/calificar', methods=['POST'])
@login_required
def calificar(id):
    ticket = Solicitud.query.get_or_404(id)
    user = Usuario.query.get(session['user_id'])
    
    if ticket.usuario_id != user.id:
        flash('Solo el solicitante puede calificar', 'danger')
        return redirect(url_for('solicitudes.ver', id=id))
    
    if ticket.estado != 'Cerrado':
        flash('Solo se pueden calificar tickets cerrados', 'danger')
        return redirect(url_for('solicitudes.ver', id=id))
    
    calificacion_valor = int(request.form.get('calificacion', 0))
    comentario = request.form.get('comentario', '')
    
    if calificacion_valor < 1 or calificacion_valor > 5:
        flash('Calificación inválida', 'danger')
        return redirect(url_for('solicitudes.ver', id=id))
    
    calificacion = CalificacionTicket.get_by_ticket(id)
    if calificacion:
        calificacion.calificacion = calificacion_valor
        calificacion.comentario = comentario
    else:
        calificacion = CalificacionTicket(
            solicitud_id=id,
            calificacion=calificacion_valor,
            comentario=comentario
        )
        db.session.add(calificacion)
    
    db.session.commit()
    flash('Gracias por tu calificación', 'success')
    
    return redirect(url_for('solicitudes.ver', id=id))

@solicitudes_bp.route('/<int:id>/quick-update', methods=['POST'])
@login_required
def quick_update(id):
    from flask import jsonify
    ticket = Solicitud.query.get_or_404(id)
    user = Usuario.query.get(session['user_id'])
    
    if not (user.is_admin or user.is_coordinador or user.is_resolutor):
        return jsonify({'error': 'Sin permisos'}), 403
    
    data = request.get_json()
    nuevo_estado = data.get('estado')
    
    if nuevo_estado in ['Planificado', 'Solucionado']:
        ticket.estado = nuevo_estado
        if nuevo_estado == 'Solucionado' and not ticket.resuelto_por:
            ticket.resuelto_por = user.id
        db.session.commit()
        return jsonify({'success': True, 'estado': nuevo_estado})
    
    return jsonify({'error': 'Estado inválido'}), 400

@solicitudes_bp.route('/<int:id>/reabrir', methods=['POST'])
@login_required
def reabrir(id):
    from models import AuditoriaTicket
    ticket = Solicitud.query.get_or_404(id)
    user = Usuario.query.get(session['user_id'])
    
    # Validar permisos: solicitante, coordinador o admin
    if not (ticket.usuario_id == user.id or user.is_coordinador or user.is_admin):
        flash('No tienes permisos para reabrir este ticket', 'danger')
        return redirect(url_for('solicitudes.ver', id=id))
    
    if ticket.estado != 'Cerrado':
        flash('Solo se pueden reabrir tickets cerrados', 'danger')
        return redirect(url_for('solicitudes.ver', id=id))
    
    motivo = request.form.get('motivo_reapertura', '').strip()
    if not motivo:
        flash('Debes proporcionar un motivo para reabrir el ticket', 'danger')
        return redirect(url_for('solicitudes.ver', id=id))
    
    # Reabrir ticket
    estado_anterior = ticket.estado
    ticket.estado = 'Pendiente'
    ticket.motivo_reapertura = motivo
    ticket.fecha_reapertura = datetime.now()
    ticket.reabierto_por = user.id
    ticket.veces_reabierto = (ticket.veces_reabierto or 0) + 1
    ticket.fecha_cierre = None
    
    # Auditoría
    auditoria = AuditoriaTicket(
        ticket_id=ticket.id,
        usuario_id=user.id,
        accion='reabrir',
        campo_modificado='estado',
        valor_anterior=estado_anterior,
        valor_nuevo='Pendiente',
        ip_address=get_client_ip(request)
    )
    db.session.add(auditoria)
    
    # Comentario automático
    comentario = ComentarioTicket(
        solicitud_id=id,
        usuario_id=user.id,
        contenido=f'🔄 Ticket reabierto por {user.nombre}. Motivo: {motivo}',
        es_interno=False
    )
    db.session.add(comentario)
    
    db.session.commit()
    flash(f'Ticket #{ticket.id} reabierto exitosamente', 'success')
    
    return redirect(url_for('solicitudes.ver', id=id))
